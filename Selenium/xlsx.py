import openpyxl
#read from excel
book=openpyxl.load_workbook("C:\\Users\\kodiak\\Desktop\\info.xlsx")
#book.create_sheet("Sheet1",0)
sheet=book.get_sheet_by_name("Sheet1")
print(sheet['A3'].value)

print(sheet.cell(row=1,column=1).value)

#write from excel
book=openpyxl.load_workbook("C:\\Users\\kodiak\\Desktop\\info.xlsx")
sheet=book.get_sheet_by_name("Sheet1")
w=sheet['A3'].value=20
book.save("C:\\Users\\kodiak\\Desktop\\info.xlsx")
print(w)
print(sheet.cell(row=1,column=1).value)
